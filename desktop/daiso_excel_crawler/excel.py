from __future__ import annotations

import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path
from typing import Any, Callable

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

LogFn = Callable[[str], None]

HEADERS = [
    "품번",
    "상품명",
    "가격",
    "브랜드",
    "평점",
    "리뷰수",
    "재고",
    "대분류",
    "중분류",
    "소분류",
    "대표이미지",
    "상품URL",
    "상세이미지URL",
    "상품정보고시/AI추출",
    "주요정보",
]


def export_products(products: list[dict[str, Any]], output_path: str | Path, log: LogFn | None = None) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "상품목록"
    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="EDEDED")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    with tempfile.TemporaryDirectory() as tmpdir:
        image_paths = download_images(products, Path(tmpdir), log=log)
        last_row = len(products) + 1
        for row_idx, product in enumerate(products, start=2):
            ws.append(row_values(product))
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row_idx, col_idx).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row_idx, 12).hyperlink = product.get("product_url")
            ws.cell(row_idx, 12).style = "Hyperlink"
            image_path = image_paths.get(row_idx)
            if image_path:
                add_image(ws, image_path, f"K{row_idx}")
                ws.row_dimensions[row_idx].height = 92
            if log and (row_idx == 2 or row_idx % 100 == 0 or row_idx == last_row):
                log(f"엑셀 행 작성: {row_idx - 1}/{len(products)}")
        widths = [14, 34, 12, 18, 10, 10, 10, 14, 14, 18, 16, 46, 70, 55, 55]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        wb.save(output)
    return output


def row_values(product: dict[str, Any]) -> list[Any]:
    ai_notice = product.get("ai_notice") or {}
    attrs = product.get("attributes_text") or ""
    notice_text = "\n".join(f"{k}: {v}" for k, v in ai_notice.items() if v) or product.get("notice_text", "")
    return [
        product.get("pd_no"),
        product.get("name"),
        product.get("price"),
        product.get("brand"),
        product.get("rating"),
        product.get("review_count"),
        product.get("stock"),
        product.get("large_category"),
        product.get("middle_category"),
        product.get("small_category"),
        product.get("image_url"),
        product.get("product_url"),
        "\n".join(product.get("detail_image_urls") or []),
        notice_text,
        attrs,
    ]


def download_images(products: list[dict[str, Any]], tmpdir: Path, log: LogFn | None = None) -> dict[int, Path]:
    image_items = [
        (row_idx, product.get("image_url") or "")
        for row_idx, product in enumerate(products, start=2)
        if product.get("image_url")
    ]
    if log:
        log(f"대표 이미지 다운로드: {len(image_items)}개")
    paths: dict[int, Path] = {}

    def task(row_idx: int, url: str) -> tuple[int, Path | None]:
        return row_idx, download_image(url, tmpdir / f"thumb_{row_idx}.jpg")

    with ThreadPoolExecutor(max_workers=16) as executor:
        futures = [executor.submit(task, row_idx, url) for row_idx, url in image_items]
        for done_count, future in enumerate(as_completed(futures), start=1):
            row_idx, image_path = future.result()
            if image_path:
                paths[row_idx] = image_path
            if log and (done_count == 1 or done_count % 250 == 0 or done_count == len(futures)):
                log(f"대표 이미지 다운로드: {done_count}/{len(futures)}")
    return paths


def download_image(url: str, output: Path) -> Path | None:
    try:
        res = requests.get(url, timeout=20, headers={"User-Agent": "Mozilla/5.0"})
        res.raise_for_status()
        with PILImage.open(BytesIO(res.content)) as img:
            img.thumbnail((110, 90))
            rgb = img.convert("RGB")
            rgb.save(output, "JPEG", quality=88)
        return output
    except Exception:
        return None


def add_image(ws, image_path: Path, cell: str) -> None:
    img = XLImage(str(image_path))
    img.anchor = cell
    ws.add_image(img)
